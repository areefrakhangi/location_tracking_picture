import subprocess
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

EXIFTOOL = r"C:\Tools\ExifTool\exiftool.exe"


def scan_folder(folder):
    cmd = [
        EXIFTOOL,
        "-r",
        "-n",
        "-FileName",
        "-Directory",
        "-FileType",
        "-DateTimeOriginal",
        "-CreateDate",
        "-MediaCreateDate",
        "-TrackCreateDate",
        "-GPSLatitude",
        "-GPSLongitude",
        "-Make",
        "-Model",
        "-ImageWidth",
        "-ImageHeight",
        "-FileSize",
        "-j",
        folder
    ]

    result = subprocess.run(cmd, capture_output=True, text=True)
    return json.loads(result.stdout)


def pick_date(item):
    return (
        item.get("DateTimeOriginal")
        or item.get("CreateDate")
        or item.get("MediaCreateDate")
        or item.get("TrackCreateDate")
        or ""
    )


def build_excel(data, output):
    wb = Workbook()
    ws = wb.active
    ws.title = "Media Report"

    headers = [
        "Folder", "File", "Type", "Date Taken",
        "Make", "Model",
        "Width", "Height",
        "Latitude", "Longitude",
        "Has GPS", "Google Maps",
        "File Size (MB)"
    ]

    ws.append(headers)

    green = PatternFill(start_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", fill_type="solid")

    total = len(data)

    for i, item in enumerate(data, start=1):

        lat = item.get("GPSLatitude", "")
        lon = item.get("GPSLongitude", "")

        has_gps = "Yes" if lat != "" and lon != "" else "No"

        maps = f"https://www.google.com/maps?q={lat},{lon}" if has_gps == "Yes" else ""

        size = item.get("FileSize", 0)
        size_mb = round(size / (1024 * 1024), 2) if isinstance(size, (int, float)) else ""

        row = [
            item.get("Directory", ""),
            item.get("FileName", ""),
            item.get("FileType", ""),
            pick_date(item),
            item.get("Make", ""),
            item.get("Model", ""),
            item.get("ImageWidth", ""),
            item.get("ImageHeight", ""),
            lat,
            lon,
            has_gps,
            maps,
            size_mb
        ]

        ws.append(row)

        fill = green if has_gps == "Yes" else red
        for cell in ws[ws.max_row]:
            cell.fill = fill

        # progress bar
        percent = (i / total) * 100
        sys.stdout.write(
            f"\rProcessed {i}/{total} files ({percent:.2f}%)"
        )
        sys.stdout.flush()

    # column sizing
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2

    ws.freeze_panes = "A2"

    wb.save(output)


def main():
    print("=== MEDIA REPORT GENERATOR ===\n")

    folder = input("Enter folder to scan: ").strip('" ')
    output = input("Output Excel file (.xlsx): ").strip('" ')

    if not output.endswith(".xlsx"):
        output += ".xlsx"

    print("\nScanning with ExifTool...\nThis may take a while...\n")

    data = scan_folder(folder)

    print(f"\n\nTotal files found: {len(data)}")
    print("Building Excel report...\n")

    build_excel(data, output)

    print("\nDONE!")
    print("Saved to:", output)


if __name__ == "__main__":
    main()